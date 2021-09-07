import openpyxl
import time
from pathlib import Path


# Deze plant moet nog worden opgedeeld in gewas en groenbemester
class Plant:
    def __init__(self, id, naam):
        self.id = id
        self.naam = naam


class Waarneming:
    def __init__(self, id, naam):
        self.id = id
        self.naam = naam


# Een aaltje is de combinatie tussen een waarneming en een plant
class Aaltje:
    def __init__(self, plant, waarneming, schade, vermeerdering, opmerking=None):
        self.plant = plant
        self.waarneming = waarneming
        self.schade = schade
        self.vermeerdering = vermeerdering
        self.opmerking = opmerking

    def zoekOfDitJuisteAaltjeIs(self, plant, waarneming):
        return self.plant == plant and self.waarneming == waarneming


class Grafiek:
    def __init__(self, waarneming):
        self.waarneming = waarneming

    def teken(self):
        aantalAaltjes = 0
        i = 0
        print(self.waarneming.naam)
        for plant in jaren:
            aaltje = self.zoekAaltje(plant)
            aantalAaltjes = self.berekenAaltjes(aaltje.vermeerdering, aantalAaltjes)
            print(i, plant.naam, aaltje.schade, aantalAaltjes)
            i = i + 1  # Waarom zit er geen i++ in python?

    def zoekAaltje(self, plant):
        for aaltje in aaltjes:
            if aaltje.zoekOfDitJuisteAaltjeIs(plant, self.waarneming):
                return aaltje
        return geenGegevensAaltje

    # Hier moet nog een of ander algoritme voor komen
    def berekenAaltjes(self, vermeerdering, startGetal):
        return startGetal + vermeerdering


def init():
    makePlanten()
    makeJaren()
    makeWaarnemingen()
    makeAaltjes()
    makeGrafieken()


def makePlanten():
    global planten
    planten = []

    xlsx_file = Path('.', 'R0004.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    i = 1
    while sheet["C" + str(i)].value != None:
        id = sheet["E" + str(i)].value
        naam = sheet["C" + str(i)].value
        plant = Plant(id, naam)
        planten.append(plant)
        i = i + 1


def makeJaren():
    global jaren
    jaren = []


def makeWaarnemingen():
    global waarnemingen
    waarnemingen = []

    xlsx_file = Path('.', 'r0043.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    waarnemingenCellen = ["185", "964", "963", "961", "959", "215"]

    for getal in waarnemingenCellen:
        waarneming = Waarneming(sheet["B" + str(getal)].value, sheet["C" + str(getal)].value)
        waarnemingen.append(waarneming)


def makeAaltjes():
    global aaltjes
    aaltjes = []

    global geenGegevensAaltje
    geenGegevensAaltje = Aaltje("Geen gegevens", "Geen gegevens", 0, 0)

    xlsx_file = Path('.', 'r0765.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    start = time.time()
    i = 1
    while sheet["C" + str(i)].value != None:
        for waarneming in waarnemingen:
            for plant in planten:
                if waarneming.id == sheet["C" + str(i)].value and plant.id == sheet["D" + str(i)].value:
                    aaltje = Aaltje(plant, waarneming, sheet["G" + str(i)].value, sheet["B" + str(i)].value)
                    aaltjes.append(aaltje)
                    break  # Deze breaks en continues zijn niet per se nodig maar maakt het iets sneller
            else:
                continue
            break
        i = i + 1
    end = time.time()
    print(end - start)
    print(len(aaltjes))


def makeGrafieken():
    global grafieken
    grafieken = []

    for waarneming in waarnemingen:
        grafiek = Grafiek(waarneming)
        grafieken.append(grafiek)


def grafiekenTekenen():
    for grafiek in grafieken:
        grafiek.teken()


def plantToevoegenAanJaren(naam):
    naamCap = naam.capitalize()
    for plant in planten:
        if plant.naam == naamCap:
            jaren.append(plant)
            return True
    return False


def printAllePlanten():
    for plant in planten:
        print(plant.naam, plant.id)


def cmd():
    print("type wat je wilt doen")
    inputStr = input()
    if inputStr == "plant toevoegen" or inputStr == "p":
        print("Welke plant wil je toevoegen?")
        inputStr = input()
        if plantToevoegenAanJaren(inputStr):
            print("Deze plant is toegevoegd")
        else:
            print("Deze plant bestaat niet")
        cmd()
    elif inputStr == "teken" or inputStr == "t":
        grafiekenTekenen()
        cmd()
    elif inputStr == "exit" or inputStr == "e":
        pass
    elif inputStr == "help" or inputStr == "h":
        print(
            "Type \"plant toevoegen\" om een plant toe te voegen, type \"teken\" om een grafiek te tekenen, type \"exit\" om af te sluiten")
        cmd()
    elif inputStr == "print alle planten" or inputStr == "pr":
        printAllePlanten()
        cmd()
    else:
        print("Dit is geen geldig commando, type \"help\" voor hulp")
        cmd()


def main():
    init()
    cmd()


if __name__ == '__main__':
    main()
